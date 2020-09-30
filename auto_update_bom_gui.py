"""
Automatically update the specified list of parts sheet based on a TC exported BoM.

"""

import tkinter as tk

from openpyxl import load_workbook, utils
from tkinter import filedialog, messagebox


class GUI:
    def __init__(self, master):
        self.master = master
        master.title("Consolidate BoM")

        self.instructions = tk.Message(master, text="Unique items are compiled from the BoM into the Part List file. "
                                                    "Quantity is summed. If a component exists in the Part List, "
                                                    "but not the BoM, it's marked in the deprecated column. "
                                                    "Only columns present in both will be updated. "
                                                    "New components are added to the end of the list.", width=300)
        self.instructions.pack()

        self.select_in_button = tk.Button(master, text="Select BoM", command=self.select_in)
        self.select_in_button.pack()

        self.select_out_button = tk.Button(master, text="Select Parts List", command=self.select_out)
        self.select_out_button.pack()

        self.consolidate_button = tk.Button(master, text="Consolidate", command=self.consolidate)
        self.consolidate_button.pack()

        self.in_wb = None
        self.out_wb = None

    def select_in(self):
        try:
            self.in_wb = load_workbook(
                filedialog.askopenfilename(title="Select BoM file",
                                           filetypes=(("excel files", "*.xls*"), ("all files", "*.*"))))
        except:
            messagebox.showinfo("", "Bad file selection. Please try again.")

    def select_out(self):
        try:
            self.out_wb = load_workbook(
                filedialog.askopenfilename(title="Select Part List file",
                                           filetypes=(("excel files", "*.xls*"), ("all files", "*.*"))))
        except:
            messagebox.showinfo("", "Bad file selection. Please try again.")

    def consolidate(self):
        """Simple program that integrates TC BoMs into excel sheets containing other information."""
        if self.in_wb is None or self.out_wb is None:
            messagebox.showinfo("No input files selected")
            return

        in_ws = self.in_wb.active
        out_ws = self.out_wb.active

        in_column_names = dict([(c.value, c.col_idx) for c in in_ws[1]])
        out_column_names = dict([(c.value, c.col_idx) for c in out_ws[1]])
        in_part_numbers = dict([(c.value, c.row) for c in in_ws[
            utils.get_column_letter(in_column_names["Item ID"])]])
        out_part_numbers = dict([(c.value, c.row) for c in out_ws[
            utils.get_column_letter(out_column_names["Item ID"])]])

        updated_this_session = []
        for in_row in in_ws.iter_rows(min_row=2):
            it_id = in_row[in_column_names["Item ID"]-1].value

            if it_id in out_part_numbers.keys():
                out_row = out_part_numbers[it_id]
            else:
                out_row = out_ws.max_row + 1
                out_part_numbers[it_id] = out_row

            for col_name in in_column_names.keys():
                if col_name in out_column_names.keys():
                    in_val = in_row[in_column_names[col_name] - 1].value
                    current_val = out_ws.cell(out_row, out_column_names[col_name]).value

                    if col_name == "Quantity":
                        if in_val is None:
                            in_val = 1
                        elif in_val == "A/R":
                            pass
                        else:
                            in_val = int(in_val)

                        if it_id not in updated_this_session:
                            new_val = in_val
                        else:
                            if in_val == "A/R" or current_val == "A/R":
                                new_val = "A/R"
                            else:
                                new_val = in_val + current_val

                    elif col_name == "Item ID":
                        new_val = in_val
                    elif in_val is not None:
                        try:
                            new_val = int(in_val)
                        except ValueError:
                            new_val = in_val
                    else:
                        new_val = None

                    out_ws.cell(out_row, out_column_names[col_name]).value = new_val

            if it_id not in updated_this_session:
                updated_this_session.append(it_id)

        for pn in out_part_numbers.keys():
            if pn not in in_part_numbers.keys():
                if "Deprecated" not in out_column_names.keys():
                    out_ws.insert_cols(1)
                    out_ws['A1'].value = "Deprecated"
                    out_column_names = dict([(c.value, c.col_idx) for c in out_ws[1]])

                dep_col = out_column_names["Deprecated"]
                out_ws.cell(out_part_numbers[pn], dep_col).value = "Yes"

        save_filename = filedialog.asksaveasfilename(title="Select output file",
                                                     filetypes=(("excel files", "*.xls*"), ("all files", "*.*")))
        if save_filename == "" or save_filename is None:
            return
        else:
            self.out_wb.save(save_filename)


if __name__ == '__main__':
    root = tk.Tk()
    gui = GUI(root)
    root.mainloop()
